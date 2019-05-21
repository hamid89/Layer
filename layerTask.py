#!/usr/bin/python
# libraries to be imported 
import numpy as np
import pandas as pd
from openpyxl import load_workbook
# in case we need to take emails of the recipients and then email can be send using python platform through different connectivity packages
def getEmail(n):
      ls = []
      for i in range(0,n):
         ls.append(raw_input('Please give your email address: '))
      return ls
# .xlsx file loading using  pandas
def dataLoad(filename,sheetName):
    data = pd.ExcelFile(filename)
    sheet = pd.read_excel(data, sheetName)
    ls = []
    for variables in sheet:
        (ls.append(variables))
    return sheet,ls

# function for selecting features (variables) from a loaded sheet  
def selections(n,data):
     lis = []
     for i in range(0,n):
          lis.append(raw_input('please enter the name of a field you want to select from above (case sensitive): '))

     return lis
# function for selecting range of cells or a single cell from selected variables from the abov function 
def cellSelection(lis,data):
   ls = []
   for i in range(len(lis)):
       templist = data[lis[i]]
       ls.append([])
       print('please follow the following instructions to select the cells for the field.. '+str(lis[i])+' !')
       temp = input('press 1 for range of cells selection and 2 for one cell selection: ')
       if (temp == 1):
           print ('Indexes should not be out of range than: '+ str(len(templist)))
           n1 = input('please Eneter the first cell index: ')
           n2 = input('please Eneter the second cell index: ')
           
           ls[i] = templist[n1:n2]
       elif (temp == 2):
             print ('Index should not be out of range than: '+ str(len(templist)))
             n = input('please enter the cell number: ')
             ls[i] = templist[n]
   return ls
#sheet saving from the selected data
def sheetSave(filename,sheetName,selectField,selectedCell):
    book = load_workbook(filename)
    sheet = book.create_sheet(sheetName)
    n = 1
    for i in range(len(selectField)):
        sheet.cell(row=1, column=n).value = selectField[i]
        n = n + 1
    m = 0
    for i in range(len(selectedCell)):
        m = m + 1
        n = 2
        for j in range(len(selectedCell[i])):
            sheet.cell(row=n, column=m).value = selectedCell[i][j]
            n = n+1
    book.save(filename)

    
fileName = raw_input('Please Enter the filepath with .extension: ')
sheetname = raw_input('please Enter the respective sheet name: ')
dataFrame,header = dataLoad(fileName,sheetname)
numberOfSelections = input('Please enter number of selctions, should be equal or less than! '+ str(len(header))+' : ')
if(numberOfSelections >= len(header)):
          numberOfSelections = input('Please enter number of selctions, should be equal or less than '+ str(len(header))+' : ')
print(header)
selectedFields = selections(numberOfSelections,dataFrame)
selectedData = cellSelection(selectedFields,dataFrame)
outputFile = raw_input('Please enter the output filename with an extension: ')
outputSheetName = raw_input('Please eneter the sheet name to store: ')
sheetSave(outputFile,outputSheetName,selectedFields,selectedData)
recipietns = input('Number of recipients: ')
emailAddresses = getEmail(recipietns) 
print(emailAddresses)




